import openpyxl, json
path = r'C:\Users\IkeFl\.openclaw\media\inbound\Quote_Master_Short_Form_2025-05-02---2e18258d-7d80-45c2-afaa-37cd76ac3209.xlsx'
wb = openpyxl.load_workbook(path, data_only=False)
out = []
for ws in wb.worksheets:
    samples = []
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 25), values_only=False):
        vals = []
        for c in row[:min(ws.max_column, 12)]:
            v = c.value
            if v is None:
                continue
            vals.append(f"{c.coordinate}={v}")
        if vals:
            samples.append(vals[:8])
    out.append({
        'sheet': ws.title,
        'rows': ws.max_row,
        'cols': ws.max_column,
        'samples': samples[:12],
    })
print(json.dumps(out, indent=2, default=str))
