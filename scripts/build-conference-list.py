import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# Colors
HDR_FILL   = PatternFill("solid", fgColor="2F5496")
OG_FILL    = PatternFill("solid", fgColor="D6E4F7")
RET_FILL   = PatternFill("solid", fgColor="E2F0D9")
GEN_FILL   = PatternFill("solid", fgColor="FFF2CC")
HDR_FONT   = Font(bold=True, color="FFFFFF")
HDR_ALIGN  = Alignment(horizontal="center", vertical="center", wrap_text=True)

conferences = [
    # Month, Show Name, Dates, Location, Vertical, Expo Pass Cost, Attendees, Notes/URL
    # ── 2025 ──
    ("Sep 2025","Shale Insight 2025","Sep 16-19, 2025","Erie, PA","Midstream / O&G","TBD","","shaleinsight.com; Marcellus Shale Coalition"),
    ("Sep 2025","GPA Midstream Annual Convention","Sep 21-24, 2025","San Antonio, TX","Midstream / O&G","TBD","1,800+","gpamidstream.org"),
    ("Sep 2025","Methane Mitigation Summit Canada","Sep 23-25, 2025","Calgary, AB","Midstream / O&G","TBD","","Industrial Decarbonization Network"),
    ("Oct 2025","SPE ATCE - Annual Technical Conference and Exhibit","Oct 6-8, 2025","Dallas, TX","Midstream / O&G","TBD","","spe.org/atce"),
    ("Oct 2025","NMOGA Annual Meeting and Exhibition","Oct 7-8, 2025","Santa Fe, NM","Midstream / O&G","TBD","","nmoga.org"),
    ("Oct 2025","Oil and Natural Gas Expo","Oct 9, 2025","Oklahoma City, OK","Midstream / O&G","TBD","","One-day expo; producers + tech providers"),
    ("Oct 2025","Retail Supply Chain and Logistics Expo 2025","Oct 15-16, 2025","Las Vegas, NV","Retail","TBD","","retailscl.com"),
    ("Oct 2025","East Texas Gas Producers Association Meeting","Oct 16, 2025","Longview, TX","Midstream / O&G","TBD","","Regional gas producers"),
    ("Nov 2025","Disasters Expo USA 2025","Nov 5-6, 2025","Houston, TX","General","TBD","",""),
    ("Nov 2025","Annual Midstream Council Meeting (OOGA)","Nov 12, 2025","Columbus, OH","Midstream / O&G","TBD","","Ohio Oil and Gas Association"),
    ("Dec 2025","9th Midstream Oil and Gas Law Conference","Dec 9, 2025","Houston, TX","Midstream / O&G","$495 regular / $440 member","","cailaw.org; early reg deadline Nov 20"),
    ("Dec 2025","Rocky Mountain Pipeliners Club Holiday Party","Dec 5, 2025","Denver, CO","Midstream / O&G","TBD","","Networking event"),
    ("Dec 2025","San Antonio Pipeliners Association Luncheon","Dec 11, 2025","San Antonio, TX","Midstream / O&G","TBD","","Networking event"),
    # ── 2026 ──
    ("Jan 2026","NRF 2026: Retail's Big Show","Jan 11-13, 2026","New York, NY","Retail","Retailers ~$100 / Non-retail $2,000-$3,500","40,000+","nrf.com - biggest retail show in the world"),
    ("Feb 2026","RILA LINK 2026 - Retail Supply Chain Conference","Feb 1-4, 2026","Orlando, FL","Retail","Complimentary for qualifying retailers","","rila.org; logistics, fulfillment, AI, automation"),
    ("Mar 2026","Houston GPA Midstream Training Session","Mar 4, 2026","Houston, TX","Midstream / O&G","TBD","","gpamidstream.org"),
    ("Mar 2026","World Hydrogen and Carbon Americas","Mar 10-12, 2026","Houston, TX","Midstream / O&G","TBD","","Hydrogen/carbon capture focus"),
    ("Mar 2026","AMPP Annual Conference + Expo","Mar 15-19, 2026","Houston, TX","Midstream / O&G","TBD","","ampp.org; pipeline corrosion/integrity"),
    ("Mar 2026","CERAWeek 2026","Mar 23-27, 2026","Houston, TX","Midstream / O&G","TBD (premium event)","5,000+","ceraweek.com - Davos of Energy"),
    ("Mar 2026","Louisiana Gas Association GASTEC Expo Day","Mar 24, 2026","Zachary, LA","Midstream / O&G","TBD","",""),
    ("Mar 2026","Shoptalk Spring 2026","Mar 24-26, 2026","Las Vegas, NV","Retail","Retailers free-$1,250 / General $3,545","8,000+","shoptalk.com"),
    ("Mar 2026","WTX Oil and Gas Convention","Mar 25-26, 2026","Midland, TX","Midstream / O&G","TBD","","Permian Basin focus"),
    ("Mar 2026","Carbon Capture Utilization and Storage (CCUS 2026)","Mar 30 - Apr 1, 2026","The Woodlands, TX","Midstream / O&G","TBD","",""),
    ("Apr 2026","Energy Workforce and Technology Council Annual Meeting","Mar 31 - Apr 1, 2026","Tucson, AZ","Midstream / O&G","TBD","",""),
    ("Apr 2026","Oklahoma Excavation Safety Expo","Apr 1-2, 2026","Oklahoma City, OK","Midstream / O&G","TBD","","Pipeline/utility safety"),
    ("Apr 2026","Permian Basin Water In Energy Conference","Apr 7-9, 2026","Midland, TX","Midstream / O&G","TBD","",""),
    ("Apr 2026","PA Independent Oil and Gas Assoc. Spring Meeting","Apr 8-9, 2026","Greensburg, PA","Midstream / O&G","TBD","","pioga.org"),
    ("Apr 2026","Retail Technology Show 2026","Apr 22-23, 2026","London, UK","Retail","TBD","","retailtechnologyshow.com"),
    ("Apr 2026","21st Pipeline Technology Conference (ptc)","Apr 27-30, 2026","Berlin, Germany","Midstream / O&G","TBD","","pipelineconf.com - global pipeline industry"),
    ("May 2026","OTC 2026 - Offshore Technology Conference","May 4-7, 2026","Houston, TX (NRG Park)","Midstream / O&G","~$50-$100 expo pass","60,000+","otcnet.org - top offshore/midstream show"),
    ("May 2026","Seamless Middle East 2026","May 12-14, 2026","Dubai, UAE","Retail","Free","","Fintech/retail/eCommerce"),
    ("May 2026","TCEQ Environmental Trade Fair and Conference","May 19-20, 2026","San Antonio, TX","Midstream / O&G","TBD","","Environmental/pipeline compliance"),
    ("May 2026","Williston Basin Petroleum Conference","May 26-30, 2026","Bismarck, ND","Midstream / O&G","TBD","","wbpc.net"),
    ("Jun 2026","Gas, LNG and The Future of Energy 2026","Jun 2-3, 2026","London, UK","Midstream / O&G","TBD","300+",""),
    ("Jun 2026","Global Energy Show Canada 2026","Jun 9-11, 2026","Calgary, AB","Midstream / O&G","TBD","","BMO Centre"),
    ("Jun 2026","Midstream Engineering and Construction Expo (EPC Show)","Jun 16-17, 2026","Houston, TX","Midstream / O&G","TBD","10,000+ / 400+ exhibitors","epcshow.com - pipelines, processing, storage, terminals"),
    ("Jun 2026","CommerceNext Growth Show 2026","Jun 24-26, 2026","New York, NY","Retail","TBD","","eCommerce/retail digital strategy"),
    ("Sep 2026","Annual Onshore Wellsite Facilities Congress","Sep 15-17, 2026","Houston, TX","Midstream / O&G","TBD","",""),
    ("Sep 2026","Subsea Pipeline Technology Congress (SPT 2026)","Sep 15-16, 2026","London, UK","Midstream / O&G","TBD","","Offshore pipeline/riser tech"),
    ("Sep 2026","GPA Midstream Convention","Sep 20-23, 2026","San Antonio, TX","Midstream / O&G","TBD","1,800+","gpamidstream.org; sponsorships open Apr 29"),
    ("Sep 2026","North Dakota Petroleum Council Annual Meeting","Sep 20-23, 2026","Watford City, ND","Midstream / O&G","TBD","","ndpc.org"),
    ("Sep 2026","International Pipeline Conference and Expo","Sep 22-24, 2026","Calgary, AB","Midstream / O&G","TBD","300+ exhibitors","Safety/digitalization/energy transition"),
    ("Sep 2026","Retail Supply Chain and Logistics Expo 2026 (New York)","Sep 30 - Oct 1, 2026","New York, NY","Retail","TBD","","retailscl.com"),
    ("Oct 2026","SPE Permian Basin Energy Conference","Oct 7-8, 2026","Midland, TX","Midstream / O&G","TBD","","spe.org"),
    ("Oct 2026","Carbon Capture, Utilization and Storage Conference","Oct 8, 2026","Houston, TX","Midstream / O&G","TBD","","InterContinental Houston"),
    ("Oct 2026","Indiana Oil and Gas Assoc. Annual Meeting and Trade Show","Oct 12-16, 2026","New Harmony, IN","Midstream / O&G","TBD","","ioga.org"),
    ("Nov 2026","ADIPEC 2026","Nov 2-5, 2026","Abu Dhabi, UAE","Midstream / O&G","TBD","160,000+","adipec.com - massive global energy show"),
    ("Nov 2026","Retail Supply Chain and Logistics Expo 2026 (London)","Nov 11-12, 2026","London, UK","Retail","TBD","","retailscl.com"),
]

wb = openpyxl.Workbook()
wb.remove(wb.active)  # remove default sheet

ALL_HEADERS = ["Month","Show Name","Dates","Location","Vertical","Expo Pass Cost","Attendees","Notes / URL"]
MONTH_HEADERS = ["Show Name","Dates","Location","Vertical","Expo Pass Cost","Attendees","Notes / URL"]

def fill_for(vertical):
    if vertical == "Midstream / O&G":
        return OG_FILL
    elif vertical == "Retail":
        return RET_FILL
    return GEN_FILL

def set_header_row(ws, headers):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = HDR_ALIGN

def autofit(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

# All Shows tab
ws_all = wb.create_sheet("All Shows")
set_header_row(ws_all, ALL_HEADERS)
for row_idx, conf in enumerate(conferences, 2):
    for col_idx, val in enumerate(conf, 1):
        c = ws_all.cell(row=row_idx, column=col_idx, value=val)
        c.fill = fill_for(conf[4])
ws_all.freeze_panes = "A2"
autofit(ws_all)

# Group by month
from collections import OrderedDict
by_month = OrderedDict()
for conf in conferences:
    month = conf[0]
    if month not in by_month:
        by_month[month] = []
    by_month[month].append(conf)

for month, rows in by_month.items():
    ws = wb.create_sheet(month)
    set_header_row(ws, MONTH_HEADERS)
    for row_idx, conf in enumerate(rows, 2):
        data = conf[1:]  # skip Month column
        for col_idx, val in enumerate(data, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.fill = fill_for(conf[4])
    ws.freeze_panes = "A2"
    autofit(ws)

out = r"C:\Users\IkeFl\.openclaw\workspace\data\conferences-2025-2026.xlsx"
wb.save(out)
print(f"Saved: {out}")
print(f"Total shows: {len(conferences)}")
