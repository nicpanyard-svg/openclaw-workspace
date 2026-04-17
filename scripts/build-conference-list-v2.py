import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

HDR_FILL   = PatternFill("solid", fgColor="2F5496")
OG_FILL    = PatternFill("solid", fgColor="D6E4F7")
RET_FILL   = PatternFill("solid", fgColor="E2F0D9")
WAT_FILL   = PatternFill("solid", fgColor="DDEBF7")
ELE_FILL   = PatternFill("solid", fgColor="FFF2CC")
REN_FILL   = PatternFill("solid", fgColor="F4CCCC")
GEN_FILL   = PatternFill("solid", fgColor="EFEFEF")
FREE_FILL  = PatternFill("solid", fgColor="C6EFCE")  # green highlight for free pass rows

HDR_FONT  = Font(bold=True, color="FFFFFF")
HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Month, Show Name, Dates, Location, Vertical, Expo Pass Cost, Free Pass?, Attendees, Notes/URL
# Free Pass? = how to get in free (exhibitor guest pass, promo code, etc.) or "" if none known
conferences = [
    # ===================== MIDSTREAM / O&G =====================
    ("Sep 2025","Shale Insight 2025","Sep 16-19, 2025","Erie, PA","Midstream / O&G","TBD","","","shaleinsight.com; Marcellus Shale Coalition"),
    ("Sep 2025","GPA Midstream Annual Convention","Sep 21-24, 2025","San Antonio, TX","Midstream / O&G","Included w/ general registration","Yes - expo (Vendor Night) included w/ reg","1,800+","gpamidstream.org; on-site reg +$100"),
    ("Sep 2025","Methane Mitigation Summit Canada","Sep 23-25, 2025","Calgary, AB","Midstream / O&G","TBD","","","Industrial Decarbonization Network"),
    ("Oct 2025","SPE ATCE - Annual Technical Conference and Exhibit","Oct 20-22, 2025","Houston, TX","Midstream / O&G","Expo free for SPE members / Full conf $675-$2,000","Yes - free expo pass for SPE members","","atce.org; George R. Brown Convention Center"),
    ("Oct 2025","NMOGA Annual Meeting and Exhibition","Oct 7-8, 2025","Santa Fe, NM","Midstream / O&G","TBD","","","nmoga.org"),
    ("Oct 2025","Oil and Natural Gas Expo","Oct 9, 2025","Oklahoma City, OK","Midstream / O&G","TBD","","","One-day expo; producers + tech providers"),
    ("Oct 2025","East Texas Gas Producers Association Meeting","Oct 16, 2025","Longview, TX","Midstream / O&G","TBD","","","Regional gas producers"),
    ("Nov 2025","Annual Midstream Council Meeting (OOGA)","Nov 12, 2025","Columbus, OH","Midstream / O&G","TBD","","","Ohio Oil and Gas Association"),
    ("Dec 2025","9th Midstream Oil and Gas Law Conference","Dec 9, 2025","Houston, TX","Midstream / O&G","$495 regular / $440 member","","","cailaw.org; early reg deadline Nov 20"),
    ("Mar 2026","Houston GPA Midstream Training Session","Mar 4, 2026","Houston, TX","Midstream / O&G","TBD","","","gpamidstream.org"),
    ("Mar 2026","World Hydrogen and Carbon Americas","Mar 10-12, 2026","Houston, TX","Midstream / O&G","TBD","","","Hydrogen/carbon capture focus"),
    ("Mar 2026","AMPP Annual Conference + Expo","Mar 15-19, 2026","Houston, TX","Midstream / O&G","1-day pass $299 / member discount avail","Yes - exhibitors distribute guest passes; ask vendors","","ampp.org; ace.ampp.org"),
    ("Mar 2026","CERAWeek 2026","Mar 23-27, 2026","Houston, TX","Midstream / O&G","Exec Pass $11,000 / Innovation Agora $3,000","","5,000+","ceraweek.com - Davos of Energy; no free option known"),
    ("Mar 2026","Louisiana Gas Association GASTEC Expo Day","Mar 24, 2026","Zachary, LA","Midstream / O&G","TBD","","",""),
    ("Mar 2026","WTX Oil and Gas Convention","Mar 25-26, 2026","Midland, TX","Midstream / O&G","TBD","","","Permian Basin focus"),
    ("Mar 2026","Carbon Capture Utilization and Storage (CCUS 2026)","Mar 30 - Apr 1, 2026","The Woodlands, TX","Midstream / O&G","TBD","","",""),
    ("Apr 2026","Energy Workforce and Technology Council Annual Meeting","Apr 1, 2026","Tucson, AZ","Midstream / O&G","TBD","","",""),
    ("Apr 2026","Oklahoma Excavation Safety Expo","Apr 1-2, 2026","Oklahoma City, OK","Midstream / O&G","TBD","","","Pipeline/utility safety"),
    ("Apr 2026","GPA Midstream Technical Conference 2026","Apr 21-23, 2026","Dallas/Plano, TX","Midstream / O&G","$525 member / $1,050 non-member (early bird)","","","gpamidstream.org; Hilton Dallas/Plano Granite Park"),
    ("Apr 2026","Permian Basin Water In Energy Conference","Apr 7-9, 2026","Midland, TX","Midstream / O&G","TBD","","",""),
    ("Apr 2026","PA Independent Oil and Gas Assoc. Spring Meeting","Apr 8-9, 2026","Greensburg, PA","Midstream / O&G","TBD","","","pioga.org"),
    ("Apr 2026","21st Pipeline Technology Conference (ptc)","Apr 27-30, 2026","Berlin, Germany","Midstream / O&G","TBD","","","pipelineconf.com - global pipeline industry"),
    ("May 2026","OTC 2026 - Offshore Technology Conference","May 4-7, 2026","Houston, TX (NRG Park)","Midstream / O&G","4-day member $295-$495 / non-member $495-$695","Yes - exhibitors offer guest passes; contact vendors exhibiting","60,000+","otcnet.org; 1-day pass also avail from $235"),
    ("May 2026","TCEQ Environmental Trade Fair and Conference","May 19-20, 2026","San Antonio, TX","Midstream / O&G","TBD","","","Environmental/pipeline compliance"),
    ("May 2026","Williston Basin Petroleum Conference","May 19-21, 2026","Bismarck, ND","Midstream / O&G","Member $600 / Non-member $650 (early bird)","","2,200+","wbpcnd.com; includes trade show + meals"),
    ("Jun 2026","Gas LNG and The Future of Energy 2026","Jun 2-3, 2026","London, UK","Midstream / O&G","TBD","","300+",""),
    ("Jun 2026","Global Energy Show Canada 2026","Jun 9-11, 2026","Calgary, AB","Midstream / O&G","TBD","","","BMO Centre"),
    ("Jun 2026","Midstream Engineering and Construction Expo (EPC Show)","Jun 16-17, 2026","Houston, TX","Midstream / O&G","TBD","Yes - exhibitors typically provide guest passes; contact vendors","10,000+ / 400+ exhibitors","epcshow.com - pipelines, processing, storage, terminals"),
    ("Sep 2026","Annual Onshore Wellsite Facilities Congress","Sep 15-17, 2026","Houston, TX","Midstream / O&G","TBD","","",""),
    ("Sep 2026","GPA Midstream Convention","Sep 20-23, 2026","San Antonio, TX","Midstream / O&G","Reg opens Jun 15, 2026 / Vendor Night booth $2,500 member","Yes - expo (Vendor Night) included w/ general reg","1,800+","gpamidstream.org; sponsorships open Apr 29"),
    ("Sep 2026","North Dakota Petroleum Council Annual Meeting","Sep 20-23, 2026","Watford City, ND","Midstream / O&G","TBD","","","ndpc.org"),
    ("Sep 2026","International Pipeline Conference and Expo","Sep 22-24, 2026","Calgary, AB","Midstream / O&G","TBD","","300+ exhibitors","Safety/digitalization/energy transition"),
    ("Oct 2026","SPE Permian Basin Energy Conference","Oct 7-8, 2026","Midland, TX","Midstream / O&G","TBD","","","spe.org"),
    ("Oct 2026","Carbon Capture Utilization and Storage Conference","Oct 8, 2026","Houston, TX","Midstream / O&G","TBD","","","InterContinental Houston"),
    ("Oct 2026","Indiana Oil and Gas Assoc. Annual Meeting and Trade Show","Oct 12-16, 2026","New Harmony, IN","Midstream / O&G","TBD","","","ioga.org"),
    ("Nov 2026","ADIPEC 2026","Nov 2-5, 2026","Abu Dhabi, UAE","Midstream / O&G","TBD","Yes - exhibitors distribute guest expo passes; massive show","160,000+","adipec.com - massive global energy show"),

    # ===================== RETAIL =====================
    ("Jan 2026","NRF 2026: Retail's Big Show","Jan 11-13, 2026","New York, NY","Retail","Retailers ~$100 / Non-retail $2,000-$3,500","Yes - retailers get free/discounted passes; tech vendors offer guest codes","40,000+","nrf.com - biggest retail show in the world"),
    ("Feb 2026","RILA LINK 2026 - Retail Supply Chain Conference","Feb 1-4, 2026","Orlando, FL","Retail","Complimentary for qualifying retailers","Yes - free for qualifying retailers/CPMs via LINK Connect program","","rila.org; logistics, fulfillment, AI, automation"),
    ("Mar 2026","Shoptalk Spring 2026","Mar 24-26, 2026","Las Vegas, NV","Retail","Retailers free-$1,250 / General $3,545","Yes - qualifying retailers (Director+) attend free","8,000+","shoptalk.com"),
    ("Apr 2026","Retail Technology Show 2026","Apr 22-23, 2026","London, UK","Retail","TBD","","","retailtechnologyshow.com"),
    ("May 2026","Seamless Middle East 2026","May 12-14, 2026","Dubai, UAE","Retail","Free","Yes - FREE to attend","","Fintech/retail/eCommerce"),
    ("Jun 2026","CommerceNext Growth Show 2026","Jun 24-26, 2026","New York, NY","Retail","TBD","","","eCommerce/retail digital strategy"),
    ("Oct 2025","Retail Supply Chain and Logistics Expo 2025","Oct 15-16, 2025","Las Vegas, NV","Retail","TBD","","","retailscl.com"),
    ("Sep 2026","Retail Supply Chain and Logistics Expo 2026 (New York)","Sep 30 - Oct 1, 2026","New York, NY","Retail","TBD","","","retailscl.com"),
    ("Nov 2026","Retail Supply Chain and Logistics Expo 2026 (London)","Nov 11-12, 2026","London, UK","Retail","TBD","","","retailscl.com"),

    # ===================== WATER =====================
    ("Mar 2025","Texas Water 2025","Mar 18-21, 2025","Houston, TX","Water","TBD","","","WEAT + Texas AWWA joint conference"),
    ("Jun 2025","AWWA ACE25 - Annual Conference and Exposition","Jun 8-11, 2025","Denver, CO","Water","TBD","","","awwa.org"),
    ("Sep 2025","AWWA Water Infrastructure Conference","Sep 14-17, 2025","Orlando, FL","Water","TBD","","","awwa.org"),
    ("Sep 2025","WEFTEC 2025","Sep 27 - Oct 1, 2025","Chicago, IL (McCormick Place)","Water","Expo Only: $140 member / $205 non-member","Yes - exhibitors offer free promo codes (e.g. Smith & Loveless, Avanti Grout)","","weftec.org; world's largest water/wastewater show"),
    ("Oct 2025","WaterSmart Innovations 2025","Oct 7-9, 2025","Reno, NV","Water","TBD","","","awwa.org"),
    ("Oct 2025","Onsite Wastewater Mega-Conference 2025","Oct 19-22, 2025","Sandusky, OH","Water","TBD","","",""),
    ("Nov 2025","AWWA Water Quality Technology Conference","Nov 9-13, 2025","TBD","Water","TBD","","","awwa.org"),
    ("Dec 2025","Partnership for Safe Water Summit","Dec 3-4, 2025","Louisville, KY","Water","TBD","","","awwa.org"),
    ("Jan 2026","MWEA Wastewater Administrators Conference (WWAdCon)","Jan 21-23, 2026","Frankenmuth, MI","Water","$420 member / $540 non-member","","","mwea.org"),
    ("Feb 2026","Pacific Water Conference 2026","Feb 10-14, 2026","Hilton Hawaiian Village, HI","Water","TBD","","","AWWA Hawaii + HWEA joint conference"),
    ("Mar 2026","AWWA/WEF Utility Management Conference","Mar 24-27, 2026","Charlotte, NC","Water","TBD","","","awwa.org + wef.org"),
    ("Apr 2026","Design-Build for Water/Wastewater Conference 2026","Apr 13-15, 2026","Grapevine, TX","Water","TBD","","",""),
    ("Apr 2026","National Water Policy Fly-In (Water Week 2026)","Apr 14-15, 2026","Washington, DC","Water","TBD","","","AWWA + WEF joint event"),
    ("Jun 2026","AWWA ACE26 - Annual Conference and Exposition","Jun 21-24, 2026","Washington, DC","Water","Expo Only: $40 utility / $185 non-utility","Yes - utility employees pay only $40; ask your water utility contact","","awwa.org; early bird ends Apr 24"),
    ("Aug 2026","AWWA Water Infrastructure Conference 2026","Aug 30 - Sep 2, 2026","Indianapolis, IN","Water","TBD","","","awwa.org"),
    ("Sep 2026","WaterPro Conference 2026","Sep 14-16, 2026","Phoenix, AZ","Water","TBD","","","National Rural Water Association; nrwa.org"),
    ("Sep 2026","WEFTEC 2026","Sep 26-30, 2026","New Orleans, LA","Water","Expo Only: $160 advance / $175 at door","Yes - exhibitors distribute free expo promo codes each year","","weftec.org; Ernest N. Morial Convention Center"),
    ("Oct 2026","WaterSmart Innovations 2026","Oct 21-23, 2026","Portland, OR","Water","TBD","","","awwa.org"),
    ("Nov 2026","AWWA Water Quality Technology Conference 2026","Nov 15-19, 2026","Memphis, TN","Water","TBD","","","awwa.org"),

    # ===================== ELECTRIC UTILITY =====================
    ("Mar 2025","DistribuTECH 2025","Mar 24-27, 2025","Dallas, TX","Electric Utility","Expo Only: ~$375 utility / ~$599 non-utility","Yes - utility employees get discounted rate; some exhibitors offer guest passes","","distributech.com; smart grid, T&D, DER"),
    ("Feb 2026","DistribuTECH 2026","Feb 2-5, 2026","San Diego, CA","Electric Utility","Expo Only: ~$375 utility / ~$599 non-utility","Yes - utility employees get discounted rate; some exhibitors offer guest passes","","distributech.com; San Diego Convention Center"),

    # ===================== RENEWABLE ENERGY =====================
    ("Jan 2025","RE+ Hawaii 2025","Jan 14-15, 2025","Honolulu, HI","Renewable Energy","Expo Only: ~$90 member / ~$110 non-member","","","re-plus.com"),
    ("Feb 2025","RE+ Northeast 2025","Feb 12-13, 2025","Boston, MA","Renewable Energy","Expo Only: $90 member / $110 non-member","","","re-plus.com"),
    ("Feb 2025","Intersolar + ESNA 2025","Feb 25-27, 2025","San Diego, CA","Renewable Energy","Expo Only: $180","Yes - CALSSA members get free expo pass; solar manufacturers offer guest codes","9,000+","iesna.com; solar + energy storage"),
    ("Mar 2025","RE+ Microgrids 2025","Mar 19-20, 2025","New Orleans, LA","Renewable Energy","Expo Only: ~$90 member / ~$110 non-member","","","re-plus.com"),
    ("Apr 2025","WindEurope Annual Event 2025","Apr 8-10, 2025","Copenhagen, Denmark","Renewable Energy","TBD","","16,200+","windeurope.org"),
    ("Apr 2025","RE+ Southeast 2025","Apr 22-23, 2025","Atlanta, GA","Renewable Energy","Expo Only: ~$90 member / ~$110 non-member","","","re-plus.com"),
    ("May 2025","CLEANPOWER 2025","May 19-21, 2025","Phoenix, AZ","Renewable Energy","TBD","","","cleanpower.org; wind, solar, storage, transmission"),
    ("May 2025","RE+ Texas 2025","May 13-14, 2025","Houston, TX","Renewable Energy","Expo Only: ~$90 member / ~$110 non-member","","","re-plus.com"),
    ("Jun 2025","RE+ Northwest 2025","Jun 11-12, 2025","Tacoma, WA","Renewable Energy","Expo Only: ~$90 member / ~$110 non-member","","","re-plus.com"),
    ("Jun 2025","RE+ Midwest 2025","Jun 9-11, 2025","Chicago, IL","Renewable Energy","TBD","","","re-plus.com"),
    ("Jun 2025","Wind Energy Science Conference (WESC) 2025","Jun 25-27, 2025","Nantes, France","Renewable Energy","TBD","","","Biennial; researchers + scientists"),
    ("Jul 2025","RE+ Mid-Atlantic 2025","Jul 10-11, 2025","Philadelphia, PA","Renewable Energy","Expo Only: ~$90 member / ~$110 non-member","","","re-plus.com"),
    ("Jul 2025","RE+ Storage 2025","Jul 31 - Aug 1, 2025","Santa Clara, CA","Renewable Energy","Expo Only: ~$90 member / ~$110 non-member","","","re-plus.com; energy storage focus"),
    ("Sep 2025","RE+ 2025 Flagship Event","Sep 8-11, 2025","Las Vegas, NV (Venetian)","Renewable Energy","Expo Only: $365 member / $515 non-member","Yes - some exhibitors (e.g. EGSA) offer free expo pass codes","1,300+ exhibitors","re-plus.com; largest clean energy event in North America"),
    ("Oct 2025","NAWEA/WindTech Conference 2025","Oct 15-17, 2025","Dallas, TX","Renewable Energy","TBD","","","North American wind energy technical conference"),
    ("Nov 2025","Solar World Congress (SWC) 2025","Nov 3-7, 2025","Fortaleza, Brazil","Renewable Energy","TBD","","","ISES official congress"),
    ("Nov 2025","New York Solar + Storage Summit 2025","Nov 12-13, 2025","New York, NY","Renewable Energy","TBD","","","NY's largest clean energy conference"),
    ("Nov 2025","Intersolar + ESNA Texas 2025","Nov 18-19, 2025","Grapevine, TX","Renewable Energy","TBD","","","iesna.com"),
    ("Feb 2026","Intersolar + ESNA 2026","Feb 18-20, 2026","San Diego, CA","Renewable Energy","Expo Only: $150 advance / $200 onsite","Yes - solar manufacturers and exhibitors offer guest codes","","iesna.com; solar + energy storage"),
    ("Apr 2026","Wood Mackenzie Solar and Energy Storage Summit 2026","Apr 29-30, 2026","Denver, CO","Renewable Energy","$1,895 standard","","450+","woodmac.com"),
    ("Apr 2026","WindEurope Annual Event 2026","Apr 21-23, 2026","Madrid, Spain (IFEMA)","Renewable Energy","Non-member ~$1,550 / Member ~$1,142 (pre-event)","Yes - students free; govt officials free; speakers free","16,000+","windeurope.org; 600+ exhibitors"),
    ("Jun 2026","CLEANPOWER 2026","Jun 1-4, 2026","Houston, TX","Renewable Energy","1-day expo ~$773 member / ~$1,145 non-member","","","cleanpower.org; George R. Brown Convention Center"),
    ("Jun 2026","Global Offshore Wind 2026","Jun 16-17, 2026","Manchester, UK","Renewable Energy","TBD","","5,000+","renewableuk.com"),
    ("Oct 2026","ASES National Solar Conference (SOLAR 2026)","Oct 19-21, 2026","Austin, TX","Renewable Energy","TBD","","","ases.org"),
    ("Oct 2026","China Wind Power 2026","Oct 14-16, 2026","Beijing, China","Renewable Energy","TBD","","120,000+","gwec.net; 1,000+ exhibitors"),
    ("Nov 2026","RE+ 2026 Flagship Event","Nov 16-19, 2026","Las Vegas Convention Center","Renewable Energy","TBD (opens May)","Yes - exhibitors offer free expo pass codes; check with solar vendors","","re-plus.com"),
    ("Nov 2026","Wind Summit Houston 2026","Nov 2026","Houston, TX","Renewable Energy","TBD","","","windsummit.org; O&M focus"),
]

wb = openpyxl.Workbook()
wb.remove(wb.active)

ALL_HEADERS   = ["Month","Show Name","Dates","Location","Vertical","Expo Pass Cost","Free Pass?","Attendees","Notes / URL"]
MONTH_HEADERS = ["Show Name","Dates","Location","Vertical","Expo Pass Cost","Free Pass?","Attendees","Notes / URL"]

def fill_for(vertical, free_pass):
    if free_pass:
        return FREE_FILL
    return {
        "Midstream / O&G": OG_FILL,
        "Retail":          RET_FILL,
        "Water":           WAT_FILL,
        "Electric Utility":ELE_FILL,
        "Renewable Energy":REN_FILL,
    }.get(vertical, GEN_FILL)

def set_header_row(ws, headers):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = HDR_ALIGN

def autofit(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 55)

# All Shows tab
ws_all = wb.create_sheet("All Shows")
set_header_row(ws_all, ALL_HEADERS)
for row_idx, conf in enumerate(conferences, 2):
    for col_idx, val in enumerate(conf, 1):
        c = ws_all.cell(row=row_idx, column=col_idx, value=val)
        c.fill = fill_for(conf[4], conf[6])
ws_all.freeze_panes = "A2"
autofit(ws_all)

# Monthly tabs
from collections import OrderedDict
by_month = OrderedDict()
for conf in conferences:
    m = conf[0]
    by_month.setdefault(m, []).append(conf)

for month, rows in by_month.items():
    ws = wb.create_sheet(month)
    set_header_row(ws, MONTH_HEADERS)
    for row_idx, conf in enumerate(rows, 2):
        data = conf[1:]
        for col_idx, val in enumerate(data, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.fill = fill_for(conf[4], conf[6])
    ws.freeze_panes = "A2"
    autofit(ws)

# Vertical tabs
verticals = ["Midstream / O&G","Retail","Water","Electric Utility","Renewable Energy"]
for vert in verticals:
    safe_title = vert.replace("/","-")[:31]
    ws = wb.create_sheet(safe_title)
    set_header_row(ws, ALL_HEADERS)
    vert_rows = [c for c in conferences if c[4] == vert]
    for row_idx, conf in enumerate(vert_rows, 2):
        for col_idx, val in enumerate(conf, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.fill = fill_for(conf[4], conf[6])
    ws.freeze_panes = "A2"
    autofit(ws)

out = r"C:\Users\IkeFl\.openclaw\workspace\data\conferences-2025-2026.xlsx"
wb.save(out)
print(f"Saved: {out}")
print(f"Total shows: {len(conferences)}")
free_pass_shows = [c for c in conferences if c[6]]
print(f"Shows with free pass options: {len(free_pass_shows)}")
for c in free_pass_shows:
    print(f"  - {c[1]}: {c[6]}")
